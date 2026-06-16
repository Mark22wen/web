"""
Step 1: 解析指标权重.xlsx，提取三套指标体系（全国/省份/地级市）
输出：
  - scheme_全国.csv
  - scheme_省份.csv
  - scheme_地级市.csv
  - excluded_indicators.csv  （无权重，暂不参与计算）
"""

import pandas as pd
import openpyxl
import os
import re

# ─── 路径配置 ──────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PUBLIC_DIR = os.path.join(BASE_DIR, '..', 'public')
OUT_DIR    = BASE_DIR   # 输出到 build_index/

WEIGHT_FILE = os.path.join(PUBLIC_DIR, '指标权重 .xlsx')   # 注意文件名有空格
SHEET_NAMES = ['全国', '省份', '地级市']   # 后三张工作表名，如有出入请修改

# ─── 工具函数 ─────────────────────────────────────────────

def read_sheet_with_ffill(filepath: str, sheet_name: str) -> pd.DataFrame:
    """
    用 openpyxl 读取工作表，把合并单元格的值向下填充，
    返回原始 DataFrame（列名用第一行）。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # 先把合并单元格区域的 top-left 值填充到整个区域
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = top_left

    data = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            if (cell.row, cell.column) in merge_map:
                row_data.append(merge_map[(cell.row, cell.column)])
            else:
                row_data.append(cell.value)
        data.append(row_data)

    df = pd.DataFrame(data)
    # 第一行作为列名，重名列自动加 _1/_2 后缀区分
    first_row = df.iloc[0]
    # strip 去掉列名首尾空格，使"权重"和"权重 "合并为同名，触发后续 _1/_2 去重
    raw_names = [str(v).strip() if pd.notna(v) else str(i) for i, v in enumerate(first_row)]
    seen = {}
    final_names = []
    for name in raw_names:
        if name in seen:
            seen[name] += 1
            final_names.append(f'{name}_{seen[name]}')
        else:
            seen[name] = 1
            final_names.append(name)
    df.columns = final_names
    df = df.iloc[1:].reset_index(drop=True)
    return df


def detect_columns(df: pd.DataFrame) -> dict:
    """
    自动识别关键列：一级、二级、三级指标、二级权重、三级权重、方向、说明/来源。
    重名列已在 read_sheet_with_ffill 里加了 _1/_2 后缀。
    返回 {role: col_name}
    """
    col_map = {}
    cols = [str(c).strip() for c in df.columns]

    # 找一级/二级/三级列
    for role, kws in [
        ('level1', ['一级指标', '一级']),
        ('level2', ['二级指标', '二级']),
        ('level3', ['三级指标', '三级', '指标名称']),
        ('direction', ['方向', '正负', '指标方向']),
        ('source',    ['说明', '来源', '原始来源', '备注']),
        ('unit',      ['单位']),
    ]:
        for kw in kws:
            matched = [c for c in cols if kw in c]
            if matched:
                col_map[role] = matched[0]
                break

    # 权重列：可能有 权重/权重_1/权重_2 等多个
    # 结构：一级 | 二级 | 权重(二级) | 三级 | 权重(三级) | ...
    # 三级指标列之后的第一个权重列才是三级权重（叶子权重）
    weight_cols = [c for c in cols if '权重' in c or c.lower() == 'weight']
    if weight_cols:
        l3_col = col_map.get('level3')
        if l3_col and l3_col in cols:
            l3_idx = cols.index(l3_col)
            # 取三级指标列之后的第一个权重列
            after_l3 = [c for c in weight_cols if cols.index(c) > l3_idx]
            if after_l3:
                col_map['weight'] = after_l3[0]       # 三级权重（叶子）
            if len(weight_cols) >= 2:
                before_l3 = [c for c in weight_cols if cols.index(c) < l3_idx]
                if before_l3:
                    col_map['weight_l2'] = before_l3[-1]  # 二级权重
        else:
            col_map['weight'] = weight_cols[-1]  # fallback 取最后一个

    # 如果三级指标列没找到，取除已识别列外的第一个非空列
    if 'level3' not in col_map:
        identified = set(col_map.values())
        for c in cols:
            if c not in identified:
                col_map['level3'] = c
                break

    return col_map


def fill_missing_weights(df: pd.DataFrame, group_cols: list, weight_col: str) -> pd.DataFrame:
    """
    同组内只缺一个权重时，用 1 - 已有权重之和 补全。
    group_cols: 分组依据，如 ['level1','level2']（三级）或 ['level1']（二级）
    weight_col: 'weight' 或 'weight_l2'
    """
    df = df.copy()
    for keys, grp in df.groupby(group_cols):
        null_mask = grp[weight_col].isna()
        n_missing = null_mask.sum()
        if n_missing == 0:
            continue
        if n_missing == 1:
            existing_sum = grp.loc[~null_mask, weight_col].sum()
            fill_val = round(1.0 - existing_sum, 6)
            if 0 < fill_val <= 1:
                idx = grp[null_mask].index[0]
                df.at[idx, weight_col] = fill_val
                print(f'    补全权重：{keys} → 缺失行填入 {fill_val:.4f}（已有权重和={existing_sum:.4f}）')
            else:
                print(f'    ⚠ {keys}：缺1个权重但已有和={existing_sum:.4f}，无法补全（结果异常），跳过')
        else:
            print(f'    ⚠ {keys}：缺 {n_missing} 个权重，超过1个无法自动补全，请手动处理')
    return df


def parse_scheme(df: pd.DataFrame, col_map: dict, sheet_name: str) -> tuple:
    """
    解析一张工作表，返回 (valid_df, excluded_df)。
    valid_df 列：level1, level2, level3, weight, direction, source, unit
    """
    records = []
    excluded = []

    l1_col   = col_map.get('level1')
    l2_col   = col_map.get('level2')
    l3_col   = col_map.get('level3')
    w_col    = col_map.get('weight')       # 三级权重（叶子）
    w2_col   = col_map.get('weight_l2')   # 二级权重
    dir_col  = col_map.get('direction')
    src_col  = col_map.get('source')
    unit_col = col_map.get('unit')

    # 对文本列做 ffill（补充合并单元格残留的空值）
    for c in [l1_col, l2_col]:
        if c and c in df.columns:
            df[c] = df[c].ffill()

    def to_float(val):
        try:
            return float(val) if val not in (None, '', 'nan', 'None') else None
        except (ValueError, TypeError):
            return None

    for _, row in df.iterrows():
        l3 = str(row.get(l3_col, '') or '').strip()
        if not l3 or l3 in ('nan', 'None', '三级指标', '指标名称', '指标'):
            continue   # 跳过标题行或空行

        w  = to_float(row.get(w_col))    # 三级权重
        w2 = to_float(row.get(w2_col))   # 二级权重（可能是 ffill 后的值）

        entry = {
            'sheet':      sheet_name,
            'level1':     str(row.get(l1_col, '') or '').strip() if l1_col else '',
            'level2':     str(row.get(l2_col, '') or '').strip() if l2_col else '',
            'level3':     l3,
            'weight':     w,    # 三级层内权重
            'weight_l2':  w2,   # 二级层内权重（按 level2 取第一次出现的值即可）
            'direction':  str(row.get(dir_col, '') or '').strip() if dir_col else '+',
            'source':     str(row.get(src_col, '') or '').strip() if src_col else '',
            'unit':       str(row.get(unit_col, '') or '').strip() if unit_col else '',
        }

        if w is None:
            excluded.append(entry)
        else:
            records.append(entry)

    return pd.DataFrame(records), pd.DataFrame(excluded)


# ─── 主流程 ───────────────────────────────────────────────

def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    all_excluded = []

    for sheet in SHEET_NAMES:
        print(f'\n══ 解析工作表：{sheet} ══')
        try:
            df_raw = read_sheet_with_ffill(WEIGHT_FILE, sheet)
        except KeyError:
            print(f'  ⚠ 未找到工作表 "{sheet}"，跳过。请检查 SHEET_NAMES 配置。')
            continue

        print(f'  原始行数：{len(df_raw)}，列：{list(df_raw.columns)}')

        col_map = detect_columns(df_raw)
        print(f'  列映射：{col_map}')

        valid_df, excl_df = parse_scheme(df_raw, col_map, sheet)
        print(f'  有效指标（初步）：{len(valid_df)} 条，排除（无权重）：{len(excl_df)} 条')

        # ── 去重：同一三级指标跨多行（说明列续行）只保留第一条，source 合并 ──
        if not valid_df.empty:
            def merge_source(sources):
                parts = [str(s).strip() for s in sources if str(s).strip() not in ('', 'nan', 'None')]
                return ' '.join(dict.fromkeys(parts))  # 去重保序拼接

            before = len(valid_df)
            valid_df = (
                valid_df.groupby(['level1', 'level2', 'level3'], sort=False)
                .agg({
                    'weight':     'first',
                    'weight_l2':  'first',
                    'direction':  'first',
                    'source':     merge_source,
                    'unit':       'first',
                    'sheet':      'first',
                })
                .reset_index()
            )
            dup = before - len(valid_df)
            if dup:
                print(f'  去重：合并了 {dup} 条重复行（同指标名跨行续写的说明已拼接）')

        if not excl_df.empty:
            excl_df = (
                excl_df.groupby(['level1', 'level2', 'level3'], sort=False)
                .agg({
                    'weight':     'first',
                    'weight_l2':  'first',
                    'direction':  'first',
                    'source':     lambda x: ' '.join(dict.fromkeys(str(s).strip() for s in x if str(s).strip() not in ('', 'nan', 'None'))),
                    'unit':       'first',
                    'sheet':      'first',
                })
                .reset_index()
            )

        # ── 自动补全缺失权重 ──
        # 三级权重：直接在行级别补全（每行一个独立三级权重）
        all_rows = pd.concat([valid_df, excl_df], ignore_index=True) if not excl_df.empty else valid_df.copy()
        print('  → 自动补全三级权重：')
        all_rows = fill_missing_weights(all_rows, ['level1', 'level2'], 'weight')
        valid_df = all_rows[all_rows['weight'].notna()].copy()
        excl_df  = all_rows[all_rows['weight'].isna()].copy()
        if not excl_df.empty:
            print(f'  补全后有效指标：{len(valid_df)} 条，仍无权重：{len(excl_df)} 条')

        # 二级权重：每个 (level1,level2) 只有一个 weight_l2，需先去重再补全再 merge 回
        if not valid_df.empty:
            print('  → 自动补全二级权重：')
            l2_unique = valid_df.drop_duplicates(subset=['level1', 'level2'])[['level1', 'level2', 'weight_l2']].copy()
            l2_filled = fill_missing_weights(l2_unique, ['level1'], 'weight_l2')
            # 把补全后的 weight_l2 merge 回 valid_df
            valid_df = valid_df.drop(columns=['weight_l2']).merge(
                l2_filled[['level1', 'level2', 'weight_l2']],
                on=['level1', 'level2'], how='left'
            )

        # 权重校验：同一二级下三级权重和应≈1
        if not valid_df.empty:
            check = valid_df.groupby(['level1', 'level2'])['weight'].sum()
            bad = check[abs(check - 1.0) > 0.05]
            if not bad.empty:
                print('  ⚠ 以下二级指标下三级权重和≠1（补全后仍有误差），请检查：')
                print(bad.to_string())

        out_path = os.path.join(OUT_DIR, f'scheme_{sheet}.csv')
        valid_df.to_csv(out_path, index=False, encoding='utf-8-sig')
        print(f'  ✓ 保存 → {out_path}')

        all_excluded.append(excl_df)

    excl_all = pd.concat(all_excluded, ignore_index=True) if all_excluded else pd.DataFrame()
    if not excl_all.empty:
        ep = os.path.join(OUT_DIR, 'excluded_indicators.csv')
        excl_all.to_csv(ep, index=False, encoding='utf-8-sig')
        print(f'\n✓ 无权重指标已保存 → {ep}')

    print('\n══ Step 1 完成 ══')
    print('请检查 scheme_*.csv 里的列映射是否正确，')
    print('如一级/二级/三级/权重列识别有误，修改 SHEET_NAMES 或 detect_columns() 里的关键词后重跑。')


if __name__ == '__main__':
    main()
