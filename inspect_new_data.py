"""
侦察脚本：分析新数据文件的结构
在 rag-backend 目录下运行：python inspect_new_data.py
需要先安装：pip install openpyxl pandas
"""
import os
import glob

# ====== 配置：把你的数据文件夹路径填在这里 ======
# 如果是在 rag-backend 根目录就用 "."，否则填绝对路径
DATA_FOLDER = r"C:\Users\DELL\Desktop\rag-backend\public"  # 改成你放数据文件的文件夹

# =============================================

def inspect_excel(path):
    import openpyxl
    print(f"\n{'='*60}")
    print(f"📊 Excel: {os.path.basename(path)}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print(f"  工作表: {wb.sheetnames}")
        for sh_name in wb.sheetnames:
            ws = wb[sh_name]
            print(f"\n  --- 工作表: {sh_name} (共 {ws.max_row} 行 × {ws.max_column} 列) ---")
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                rows.append(row)
                if len(rows) >= 4:
                    break
            for r in rows:
                cells = [str(c)[:25] if c is not None else '' for c in (r[:12] if r else [])]
                print("  | " + " | ".join(cells))
    except Exception as e:
        print(f"  ❌ 读取失败: {e}")

def list_files():
    print(f"\n📁 扫描文件夹: {DATA_FOLDER}")

    for ext in ['*.xlsx', '*.xls']:
        for f in glob.glob(os.path.join(DATA_FOLDER, '**', ext), recursive=True):
            if 'node_modules' in f or '.git' in f:
                continue
            inspect_excel(f)

    print(f"\n{'='*60}")
    print("📄 PDF 文件：")
    for f in glob.glob(os.path.join(DATA_FOLDER, '**', '*.pdf'), recursive=True):
        if 'node_modules' not in f:
            print(f"  {f}")

    print(f"\n{'='*60}")
    print("📝 Word 文件：")
    for f in glob.glob(os.path.join(DATA_FOLDER, '**', '*.docx'), recursive=True):
        if 'node_modules' not in f:
            print(f"  {f}")

if __name__ == '__main__':
    list_files()
    print("\n\n✅ 侦察完成！把上面的输出发给助手。")
